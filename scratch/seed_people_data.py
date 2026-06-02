import openpyxl
import urllib.request
import json
import sys
from datetime import datetime, date

# Supabase Configurations
SUPABASE_URL = "https://ngtjhwswbbivqajtpjvg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5ndGpod3N3YmJpdnFhanRwanZnIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2OTE4MjM2MCwiZXhwIjoyMDg0NzU4MzYwfQ.2TPnOfnAzeWG23Y-VuDKxxzQ9QdbHwrnHdVBhS9hU28"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

def http(method, path, body=None):
    url = SUPABASE_URL + path
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            text = r.read().decode()
            return r.status, json.loads(text) if text else []
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        return e.code, err

# DZM Employees list for company detection
DZM_LIST = [
    "christopher da silva souza gabassi",
    "paulo mesquita da cunha",
    "olavo carvalho de jesus montte carl",
    "victor luis de oliveira silva",
    "pedro enzo martins bastos",
    "andrea aparecida de oliveira ferreira",
    "joel maicon cardoso de oliveira",
    "robson roque bernardo",
    "wanderley souza da silva",
    "deuzelina almeida e silva",
    "flavio pereira da silva",
    "elivando santos batista",
    "joao marcelo de oliveira paredes",
    "matheus roberto santos da silva",
    "edney osmar pires",
    "rafael andrade de jesus",
    "jhonatan silva barbosa"
]

def clean_name(name):
    if not name:
        return ""
    # Normalize name: strip, lowercase, replace double spaces, strip accents if possible
    # But simple lower + strip + replace multiple spaces with single space is extremely robust
    n = str(name).strip().lower()
    return " ".join(n.split())

def parse_date(val):
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    # Try parsing string
    s = str(val).strip().split(" ")[0]
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except:
        pass
    try:
        dt = datetime.strptime(s, "%d/%m/%Y")
        return dt.strftime("%Y-%m-%d")
    except:
        pass
    return None

def clean_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("="):
        # Formula string, since we load with data_only=True it shouldn't happen, but let's be safe
        return 0.0
    s = s.replace("R$", "").replace(".", "").replace(",", ".").replace(" ", "").strip()
    try:
        return float(s)
    except:
        return 0.0

def main():
    print("=" * 80)
    print("SEEDER: Dianna.xlsx -> Supabase DB Tables")
    print("=" * 80)

    # 1. Fetch existing employees from DB
    print("\n[1/5] Buscando colaboradores existentes no banco...")
    status, db_employees = http("GET", "/rest/v1/employees?select=id,full_name,company,employment_type")
    if status != 200:
        print(f"[ERROR] Falha ao buscar colaboradores: {db_employees}")
        sys.exit(1)
    
    print(f"Encontrados {len(db_employees)} colaboradores no banco de dados.")
    
    # Create a map of cleaned name -> id
    emp_map = {}
    for emp in db_employees:
        cleaned = clean_name(emp['full_name'])
        emp_map[cleaned] = emp['id']
        
    print("\n[2/5] Carregando Dianna.xlsx...")
    wb = openpyxl.load_workbook("D:/DRE-V33-Dianna/dashboard-v2/public/Dianna.xlsx", data_only=True)
    print("Planilha carregada com sucesso!")

    # 3. Process CONSOLIDADO ATIVO INATIVO NOVO
    print("\n[3/5] Processando Trajetórias/Vínculos de CONSOLIDADO ATIVO INATIVO NOVO...")
    ws = wb['CONSOLIDADO ATIVO INATIVO NOVO']
    rows = list(ws.iter_rows(values_only=True))
    
    bonds_to_insert = []
    
    for idx, row in enumerate(rows[1:]):
        if not row[1] or not str(row[1]).strip():
            continue
        
        name = str(row[1]).strip()
        cleaned = clean_name(name)
        
        # Determine company
        company = "MarBR"
        if cleaned in DZM_LIST or "dzm" in cleaned or "rancho" in clean_name(row[2]):
            company = "DZM"
            
        emp_id = emp_map.get(cleaned)
        
        if not emp_id:
            # Create missing employee
            tipo = str(row[6]).strip() if row[6] else "CLT"
            if tipo not in ["CLT", "MEI", "Estagiário", "PJ"]:
                tipo = "CLT"
            
            status_rh = "Ativo" if "Ativo" in str(row[8]) else "Inativo"
            rem = clean_float(row[13]) if row[13] else 0.0
            st_date = parse_date(row[9])
            
            payload = {
                "full_name": name,
                "company": company,
                "employment_type": "PJ" if tipo == "MEI" else tipo,
                "status": status_rh,
                "remuneration": rem,
                "start_date": st_date,
                "job_role": str(row[4]).strip() if row[4] else None,
                "department": str(row[2]).strip() if row[2] else None,
                "active": status_rh == "Ativo"
            }
            
            print(f"  [NEW] Cadastrando novo funcionario: {name} ({company} - {tipo})")
            ins_status, ins_res = http("POST", "/rest/v1/employees", [payload])
            if ins_status in (200, 201) and ins_res:
                emp_id = ins_res[0]['id']
                emp_map[cleaned] = emp_id
            else:
                print(f"  [ERROR] Erro ao cadastrar {name}: {ins_res}")
                continue

        # Prepare employment bond
        tipo = str(row[6]).strip() if row[6] else "CLT"
        if tipo not in ["CLT", "MEI", "Estagiário", "PJ"]:
            tipo = "CLT"
            
        st_date = parse_date(row[9])
        end_date = parse_date(row[10])
        status_rh = str(row[8]).strip() if row[8] else "Ativo"
        
        if st_date:
            bonds_to_insert.append({
                "employee_id": emp_id,
                "vinculo": tipo,
                "empresa": company,
                "centro_custo": str(row[5]).strip() if row[5] else None,
                "setor": str(row[2]).strip() if row[2] else None,
                "cargo": str(row[4]).strip() if row[4] else None,
                "start_date": st_date,
                "end_date": end_date,
                "motivo_fim": "Desligamento" if end_date else None
            })

    # Insert Bonds in Supabase
    print(f"  Total de vínculos preparados: {len(bonds_to_insert)}")
    if bonds_to_insert:
        # Clear existing bonds first to prevent duplicate seeds
        http("DELETE", "/rest/v1/people_employment_bonds?id=neq.00000000-0000-0000-0000-000000000000")
        ins_status, ins_res = http("POST", "/rest/v1/people_employment_bonds", bonds_to_insert)
        if ins_status in (200, 201):
            print(f"  [SUCCESS] Trajetorias gravadas com sucesso: {len(bonds_to_insert)} registros!")
        else:
            print(f"  [ERROR] Erro ao gravar trajetorias: {ins_res}")

    # 4. Parse Monthly costs
    print("\n[4/5] Processando Lançamentos Mensais (MEI - NOVA)...")
    ws_mei = wb['MEI - NOVA']
    rows_mei = list(ws_mei.iter_rows(values_only=True))
    header_mei = rows_mei[0]
    
    # Identify month columns
    month_cols_mei = []
    for col_idx, h in enumerate(header_mei):
        if h and isinstance(h, (datetime, date)) or (h and "-" in str(h) and len(str(h)) >= 10):
            d_str = parse_date(h)
            if d_str:
                month_cols_mei.append((col_idx, d_str))
                
    print(f"  Meses identificados no MEI ({len(month_cols_mei)}): {[m[1] for m in month_cols_mei]}")
    
    costs_to_insert = []
    
    for row in rows_mei[1:]:
        if not row[1] or not str(row[1]).strip():
            continue
        
        name = str(row[1]).strip()
        cleaned = clean_name(name)
        emp_id = emp_map.get(cleaned)
        
        if not emp_id:
            continue
            
        for col_idx, competency in month_cols_mei:
            val = row[col_idx]
            val_float = clean_float(val)
            if val_float > 0:
                costs_to_insert.append({
                    "employee_id": emp_id,
                    "competencia": competency,
                    "vinculo_tipo": "MEI",
                    "valor_liquido": val_float,
                    "origem": "dianna_import",
                    "observacao": "Importado do consolidado MEI"
                })

    print("\n[5/5] Processando Lançamentos Mensais (CLT NOVA)...")
    ws_clt = wb['CLT NOVA']
    rows_clt = list(ws_clt.iter_rows(values_only=True))
    header_clt = rows_clt[0]
    
    # Identify month columns
    month_cols_clt = []
    for col_idx, h in enumerate(header_clt):
        if h and isinstance(h, (datetime, date)) or (h and "-" in str(h) and len(str(h)) >= 10):
            d_str = parse_date(h)
            if d_str:
                month_cols_clt.append((col_idx, d_str))
                
    print(f"  Meses identificados no CLT ({len(month_cols_clt)}): {[m[1] for m in month_cols_clt]}")
    
    # Parse in repeating block logic
    block = []
    current_emp_id = None
    current_emp_name = None
    
    for row in rows_clt[1:]:
        if not any(c is not None for c in row):
            continue
            
        # Is it a header row? (First cell represents employee list index)
        first_cell = str(row[0]).strip() if row[0] is not None else ""
        if first_cell.isdigit() and row[1]:
            # Flush previous block if exists
            if current_emp_id and block:
                process_clt_block(current_emp_id, block, month_cols_clt, costs_to_insert)
            
            # Start new block
            current_emp_name = str(row[1]).strip()
            current_emp_id = emp_map.get(clean_name(current_emp_name))
            block = []
        else:
            if current_emp_id:
                block.append(row)
                
    # Flush last block
    if current_emp_id and block:
        process_clt_block(current_emp_id, block, month_cols_clt, costs_to_insert)

    # Insert costs in Supabase
    print(f"\n  Total de lançamentos de custos preparados: {len(costs_to_insert)}")
    if costs_to_insert:
        # Clear existing costs first to prevent duplicate seeds
        http("DELETE", "/rest/v1/people_monthly_costs?id=neq.00000000-0000-0000-0000-000000000000")
        
        # Batch insert to prevent HTTP payload overflow (inserts 200 at a time)
        batch_size = 200
        total_inserted = 0
        for i in range(0, len(costs_to_insert), batch_size):
            batch = costs_to_insert[i:i+batch_size]
            ins_status, ins_res = http("POST", "/rest/v1/people_monthly_costs", batch)
            if ins_status in (200, 201):
                total_inserted += len(batch)
            else:
                print(f"  [ERROR] Erro ao gravar lote de custos: {ins_res}")
                
        print(f"  [SUCCESS] Custos mensais historicos gravados com sucesso: {total_inserted} registros!")

    print("\n" + "=" * 80)
    print("PROCESSO DE POPULAÇÃO DO BANCO DE DADOS CONCLUÍDO!")
    print("=" * 80)

def process_clt_block(emp_id, block, month_cols, costs_list):
    # Map variable rows in this employee's block
    row_mapping = {}
    
    # We want to identify the row type by the label in Col A (row[0])
    for r in block:
        label = clean_name(r[0]) if r[0] else ""
        if not label:
            continue
            
        if "holerite" in label:
            row_mapping["holerite"] = r
        elif "adiantamento" in label:
            row_mapping["adiantamento"] = r
        elif "hora extra" in label:
            row_mapping["hora_extra"] = r
        elif "adicional noturno" in label:
            row_mapping["adicional_not"] = r
        elif "vr" in label:
            row_mapping["vr"] = r
        elif "vt" in label:
            row_mapping["vt"] = r
        elif "ajuda de custo" in label:
            row_mapping["ajuda_custo"] = r
        elif "cesta" in label:
            row_mapping["cesta"] = r
        elif "férias" in label:
            row_mapping["ferias"] = r
        elif "rescisão" in label:
            row_mapping["rescisao"] = r
        elif "13º" in label or "13o" in label:
            row_mapping["13"] = r
        elif "descontos" in label:
            row_mapping["descontos"] = r
        elif "sum" in label or "total" in label or "-j16" in label:
            row_mapping["net"] = r

    # If we didn't find specific row indicators, the last row is typically the net liquid!
    if "net" not in row_mapping and block:
        row_mapping["net"] = block[-1]

    # For each month competency
    for col_idx, competency in month_cols:
        # Pull values
        h = clean_float(row_mapping.get("holerite")[col_idx]) if "holerite" in row_mapping else 0.0
        ad = clean_float(row_mapping.get("adiantamento")[col_idx]) if "adiantamento" in row_mapping else 0.0
        he = clean_float(row_mapping.get("hora_extra")[col_idx]) if "hora_extra" in row_mapping else 0.0
        an = clean_float(row_mapping.get("adicional_not")[col_idx]) if "adicional_not" in row_mapping else 0.0
        vr = clean_float(row_mapping.get("vr")[col_idx]) if "vr" in row_mapping else 0.0
        vt = clean_float(row_mapping.get("vt")[col_idx]) if "vt" in row_mapping else 0.0
        ac = clean_float(row_mapping.get("ajuda_custo")[col_idx]) if "ajuda_custo" in row_mapping else 0.0
        cest = clean_float(row_mapping.get("cesta")[col_idx]) if "cesta" in row_mapping else 0.0
        fer = clean_float(row_mapping.get("ferias")[col_idx]) if "ferias" in row_mapping else 0.0
        resc = clean_float(row_mapping.get("rescisao")[col_idx]) if "rescisao" in row_mapping else 0.0
        dec = clean_float(row_mapping.get("13")[col_idx]) if "13" in row_mapping else 0.0
        desc = clean_float(row_mapping.get("descontos")[col_idx]) if "descontos" in row_mapping else 0.0
        net = clean_float(row_mapping.get("net")[col_idx]) if "net" in row_mapping else 0.0
        
        # Calculate calculated net if raw net is not found or is 0
        calculated_net = (h + ad + he + an + vr + vt + ac + cest + fer + resc + dec) - desc
        final_net = net if net > 0 else calculated_net

        # Skip if nothing is paid this month
        if final_net <= 0 and h <= 0:
            continue
            
        costs_list.append({
            "employee_id": emp_id,
            "competencia": competency,
            "vinculo_tipo": "CLT",
            "valor_holerite": h,
            "valor_adiantamento": ad,
            "valor_hora_extra": he,
            "valor_adicional_not": an,
            "valor_vr": vr,
            "valor_vt": vt,
            "valor_ajuda_custo": ac,
            "valor_cesta": cest,
            "valor_ferias": fer,
            "valor_rescisao": resc,
            "valor_decimo_terceiro": dec,
            "valor_descontos": desc,
            "valor_liquido": final_net,
            "origem": "dianna_import",
            "observacao": "Importado do bloco mensal CLT"
        })

if __name__ == "__main__":
    main()
