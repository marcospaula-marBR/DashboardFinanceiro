import openpyxl

wb = openpyxl.load_workbook('D:/DRE-V33-Dianna/dashboard-v2/public/Dianna.xlsx', read_only=True)

# 1. Inspect CLT NOVA
ws = wb['CLT NOVA']
print("=== CLT NOVA ROWS ===")
for idx, r in enumerate(ws.iter_rows(values_only=True)):
    if idx < 30:
        if any(c is not None for c in r):
            print(f"Row {idx}: {[str(c)[:40] for c in r if c is not None][:15]}")

# 2. Inspect MEI - NOVA
ws2 = wb['MEI - NOVA']
print("\n=== MEI - NOVA ROWS ===")
for idx, r in enumerate(ws2.iter_rows(values_only=True)):
    if idx < 20:
        if any(c is not None for c in r):
            print(f"Row {idx}: {[str(c)[:40] for c in r if c is not None][:15]}")
