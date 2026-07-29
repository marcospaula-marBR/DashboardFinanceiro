import os
import json
import datetime
import openpyxl

def format_currency(value):
    if value is None or value == '' or value == '-':
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    try:
        s = str(value).replace('R$', '').strip()
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        val = float(s)
        return round(val, 2)
    except:
        return 0.0

def parse_date_val(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d')
    if val:
        s = str(val).strip()
        if '/' in s:
            parts = s.split('/')
            if len(parts) == 3:
                d, m, y = parts
                if len(y) == 2: y = '20' + y
                return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return None

def parse_month_col(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-01')
    if not val:
        return None
    s = str(val).strip().lower()
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 2:
            m_map = {'jan': '01', 'fev': '02', 'mar': '03', 'abr': '04', 'mai': '05', 'jun': '06',
                     'jul': '07', 'ago': '08', 'set': '09', 'out': '10', 'nov': '11', 'dez': '12'}
            m = m_map.get(parts[0][:3])
            y = "20" + parts[1] if len(parts[1]) == 2 else parts[1]
            if m and y.isdigit():
                return f"{y}-{m}-01"
    return None

def process_flat_records(sheet, sheet_name):
    records = []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return records
    
    header = rows[0]
    month_cols = {}
    for idx, cell in enumerate(header):
        parsed = parse_month_col(cell)
        if parsed:
            month_cols[idx] = parsed
            
    name_col_idx = 1
    for idx, cell in enumerate(header[:10]):
        c_str = str(cell or '').upper()
        if 'FUNCION' in c_str or 'PRESTADOR' in c_str or 'NOME' in c_str:
            name_col_idx = idx
            break

    current_nome = None
    for row in rows[1:]:
        if not row: continue
        raw_name = row[name_col_idx] if len(row) > name_col_idx else None
        if raw_name and str(raw_name).strip() and str(raw_name).strip().upper() not in ["TOTAL", "MÉDIA", "NONE", "ITEM"]:
            current_nome = str(raw_name).strip()
            
        if not current_nome: continue
        
        for col_idx, comp_date in month_cols.items():
            if col_idx >= len(row): continue
            cell_val = row[col_idx]
            if cell_val is None or cell_val == '-' or cell_val == '': continue
            val = format_currency(cell_val)
            if val <= 0: continue
            
            records.append({
                "sheet": sheet_name,
                "competencia": comp_date,
                "nome_bruto": current_nome,
                "valor_total": val
            })
            
    return records

def process_structured_clt(sheet):
    employees = []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows: return employees
    
    header = rows[0]
    header_months = {}
    for idx, c in enumerate(header[9:], start=9):
        parsed = parse_month_col(c)
        if parsed:
            header_months[idx] = parsed
            
    i = 0
    while i < len(rows):
        row = rows[i]
        if row and str(row[0]).strip().upper() == 'ITEM':
            if i + 1 < len(rows):
                emp_row = rows[i+1]
                nome = str(emp_row[1] or '').strip()
                if nome and nome.upper() not in ['TOTAL', 'MÉDIA', 'NONE', 'FUNCIONÁRIOS']:
                    status = str(emp_row[2] or '').strip()
                    setor = str(emp_row[3] or '').strip()
                    cargo_in = str(emp_row[4] or '').strip()
                    cargo_ult = str(emp_row[5] or '').strip()
                    data_in = parse_date_val(emp_row[6])
                    data_desl = parse_date_val(emp_row[7])
                    
                    monthly_data = {}
                    for r_idx in range(i+1, min(i+16, len(rows))):
                        v_row = rows[r_idx]
                        if not v_row: continue
                        verba_name = str(v_row[8] or '').strip()
                        if not verba_name or verba_name.upper() == 'ITEM': break
                        
                        for col_idx, comp_date in header_months.items():
                            if col_idx < len(v_row):
                                val = format_currency(v_row[col_idx])
                                if val > 0:
                                    if comp_date not in monthly_data:
                                        monthly_data[comp_date] = {}
                                    monthly_data[comp_date][verba_name] = val
                    
                    if monthly_data or status or setor or cargo_ult:
                        employees.append({
                            "nome": nome,
                            "status": status or 'Ativo',
                            "setor": setor or 'Operacional',
                            "cargo_inicial": cargo_in or cargo_ult or 'Não especificado',
                            "ultimo_cargo": cargo_ult or cargo_in or 'Não especificado',
                            "data_admissao": data_in,
                            "data_desligamento": data_desl,
                            "competencias": monthly_data,
                            "tipo_vinculo": "CLT"
                        })
        i += 1
    return employees

def main():
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'dashboard-v2', 'public', 'Dianna.xlsx')
    if not os.path.exists(excel_path):
        print("Excel não encontrado:", excel_path)
        return
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_flat = []
    
    target_sheets = ['MEI - NOVA', 'CLT NOVA', 'Rescisório - CLT']
    for s_name in target_sheets:
        if s_name in wb.sheetnames:
            recs = process_flat_records(wb[s_name], s_name)
            print(f"[{s_name}] Extraídos {len(recs)} lançamentos flat.")
            all_flat.extend(recs)
            
    unique_flat = []
    seen = set()
    for r in all_flat:
        key = (r['sheet'], r['competencia'], r['nome_bruto'].lower(), r['valor_total'])
        if key not in seen:
            seen.add(key)
            unique_flat.append(r)

    structured_clt = []
    if 'CLT NOVA' in wb.sheetnames:
        structured_clt = process_structured_clt(wb['CLT NOVA'])
        print(f"[CLT NOVA Structured] Extraídos {len(structured_clt)} colaboradores completos com perfil e verbas.")
            
    output_payload = {
        "records": unique_flat,
        "structured_clt": structured_clt,
        "meta": {
            "total_records": len(unique_flat),
            "total_structured_clt": len(structured_clt),
            "generated_at": datetime.datetime.now().isoformat()
        }
    }
    
    output_path = os.path.join(os.path.dirname(__file__), '..', 'dashboard-v2', 'public', 'dianna_source.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_payload, f, ensure_ascii=False, indent=2)
        
    print(f"\nData Lake Dianna v2 Gerado com sucesso!")
    print(f"Total Flat Records: {len(unique_flat)}")
    print(f"Total Structured CLT Employees: {len(structured_clt)}")
    print(f"Salvo em: {output_path}")

if __name__ == '__main__':
    main()
