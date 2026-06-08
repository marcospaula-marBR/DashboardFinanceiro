import os
import json
import re
import openpyxl

def format_currency(value):
    try:
        val = float(str(value).replace(',', '.').replace('R$', '').strip())
        return round(val, 2)
    except:
        return 0.0

def main():
    print("Iniciando geração do Data Lake Dianna...")
    
    wb_path = os.path.join(os.path.dirname(__file__), '..', 'dashboard-v2', 'public', 'Dianna.xlsx')
    if not os.path.exists(wb_path):
        print(f"Erro: Planilha não encontrada em {wb_path}")
        return
        
    print(f"Lendo {wb_path}...")
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    
    data_lake = []
    
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        
        # Tentar extrair a competência do nome da aba
        comp_match = re.search(r'(\d{1,2})[-/](\d{4})', sheet_name)
        if comp_match:
            month, year = comp_match.groups()
            competencia = f"{year}-{month.zfill(2)}-01"
        else:
            competencia = "2024-01-01"
            
        rows = list(sh.iter_rows(values_only=True))
        if not rows: continue
        
        header_row_idx = -1
        name_col = -1
        total_col = -1
        
        # Identificar cabeçalhos
        for i, row in enumerate(rows[:10]):
            row_strs = [str(cell).upper() if cell else "" for cell in row]
            for j, cell in enumerate(row_strs):
                if "COLAB" in cell or "NOME" in cell or "PRESTADOR" in cell or "FUNCIONÁRIO" in cell or "FUNCIONARIO" in cell:
                    header_row_idx = i
                    name_col = j
                    break
            if header_row_idx != -1:
                for k, c2 in enumerate(row_strs):
                    if "TOTAL" in c2 or "VALOR PAGO" in c2 or "LÍQUIDO" in c2 or "LIQUIDO" in c2:
                        total_col = k
                        break
                break
                
        if header_row_idx == -1 or name_col == -1:
            continue
            
        print(f"Processando aba: {sheet_name} (Nome na col {name_col}, Total na col {total_col})")
        
        for row in rows[header_row_idx+1:]:
            if not row or len(row) <= name_col: continue
            
            nome_bruto = str(row[name_col]).strip() if row[name_col] else ""
            if not nome_bruto or nome_bruto.upper() in ["NONE", "TOTAL", "MÉDIA", "SOMA"]: continue
            
            total_val = 0.0
            if total_col != -1 and len(row) > total_col:
                total_val = format_currency(row[total_col])
                
            if total_val == 0.0:
                continue # Pular linhas vazias
                
            data_lake.append({
                "sheet": sheet_name,
                "competencia": competencia,
                "nome_bruto": nome_bruto,
                "valor_total": total_val
            })

    output_file = os.path.join(os.path.dirname(__file__), '..', 'dashboard-v2', 'public', 'dianna_source.json')
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data_lake, f, ensure_ascii=False, indent=2)

    print(f"\nData Lake gerado com sucesso!")
    print(f"[{len(data_lake)}] registros encontrados.")
    print(f"Arquivo salvo em: {output_file}")

if __name__ == "__main__":
    main()
