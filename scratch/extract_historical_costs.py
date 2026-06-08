import os
import sys
import csv
import re
from datetime import datetime
from difflib import SequenceMatcher
import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

def similar(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def format_currency(value):
    try:
        val = float(str(value).replace(',', '.').replace('R$', '').strip())
        return round(val, 2)
    except:
        return 0.0

def main():
    print("Iniciando extração do Histórico de Custos da Planilha Dianna...")
    
    # 1. Conectar ao Supabase
    env_path = os.path.join(os.path.dirname(__file__), '..', 'dashboard-v2', '.env')
    load_dotenv(env_path)
    url: str = os.environ.get("SUPABASE_URL")
    key: str = os.environ.get("SUPABASE_SERVICE_KEY")
    
    if not url or not key:
        print("Erro: SUPABASE_URL ou SERVICE_ROLE_KEY não encontrados.")
        sys.exit(1)

    supabase: Client = create_client(url, key)
    
    # 2. Buscar colaboradores ativos no DB
    print("Buscando colaboradores no banco de dados...")
    response = supabase.table("employees").select("id, full_name, remuneration, remuneration_fixed, remuneration_bonus, remuneration_commission").eq("status", "Ativo").execute()
    db_employees = response.data
    print(f"[{len(db_employees)}] colaboradores ativos encontrados no Supabase.")

    # 3. Ler Planilha Dianna
    wb_path = os.path.join(os.path.dirname(__file__), '..', 'dashboard-v2', 'public', 'Dianna.xlsx')
    if not os.path.exists(wb_path):
        print(f"Erro: Planilha não encontrada em {wb_path}")
        sys.exit(1)
        
    print(f"Lendo {wb_path}...")
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    
    # Procurar abas de MENSAL LÍQUIDO
    target_sheets = [s for s in wb.sheetnames if 'LÍQUIDO' in s.upper() or 'MENSAL' in s.upper()]
    if not target_sheets:
        print("Aba de 'MENSAL LÍQUIDO' não encontrada. Verifique o nome da aba.")
        sys.exit(1)

    output_file = os.path.join(os.path.dirname(__file__), 'Custos_Extraidos_Revisao.csv')
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(['Aba_Origem', 'Colaborador_Planilha', 'ID_Supabase', 'Nome_Supabase', 'Competencia_YYYY_MM_01', 'Valor_Base', 'Valor_Bonus', 'Comissao', 'Descontos_Glosas', 'Valor_Liquido', 'Observacoes'])
        
        for sheet_name in target_sheets:
            print(f"Processando aba: {sheet_name}")
            sh = wb[sheet_name]
            
            # Tentar extrair a competência do nome da aba ou assumir o mês atual (apenas para template)
            # Ex: "Mensal Líquido Colaboradores Mar Brasil 09/2022" -> "2022-09-01"
            comp_match = re.search(r'(\d{1,2})/(\d{4})', sheet_name)
            if comp_match:
                month, year = comp_match.groups()
                competencia = f"{year}-{month.zfill(2)}-01"
            else:
                competencia = "2024-01-01" # Default placeholder
            
            rows = list(sh.iter_rows(values_only=True))
            if not rows: continue
            
            # Procurar linha de cabeçalho
            header_row_idx = -1
            name_col = -1
            total_col = -1
            
            for i, row in enumerate(rows[:10]):
                row_strs = [str(cell).upper() if cell else "" for cell in row]
                for j, cell in enumerate(row_strs):
                    if "COLAB" in cell or "NOME" in cell or "PRESTADOR" in cell:
                        header_row_idx = i
                        name_col = j
                        break
                if header_row_idx != -1:
                    # Tentar achar a coluna de Total
                    for k, c2 in enumerate(row_strs):
                        if "TOTAL" in c2:
                            total_col = k
                            break
                    break
            
            if header_row_idx == -1 or name_col == -1:
                print(f"  Não foi possível identificar o cabeçalho na aba {sheet_name}. Ignorando.")
                continue
                
            print(f"  Cabeçalho encontrado na linha {header_row_idx+1}. Coluna Nome: {name_col}, Coluna Total: {total_col}")
            
            for row in rows[header_row_idx+1:]:
                if not row or len(row) <= name_col: continue
                
                sheet_name_val = str(row[name_col]).strip() if row[name_col] else ""
                if not sheet_name_val or sheet_name_val.upper() in ["NONE", "TOTAL", "MÉDIA"]: continue
                
                total_val = 0.0
                if total_col != -1 and len(row) > total_col:
                    total_val = format_currency(row[total_col])
                
                # Match com Supabase
                best_match = None
                best_score = 0
                for db_emp in db_employees:
                    score = similar(sheet_name_val, db_emp['full_name'])
                    if score > best_score:
                        best_score = score
                        best_match = db_emp
                
                if best_match and best_score > 0.6: # Limiar razoável de similaridade
                    # Neste script inicial, jogamos o valor base como o total e 0 pros resto, 
                    # para o RH validar. Pode-se mapear as outras colunas (Ajuda custo, bônus, etc) se souber os índices exatos!
                    
                    db_id = best_match['id']
                    db_name = best_match['full_name']
                    base = total_val
                    bonus = 0.0
                    comissao = 0.0
                    descontos = 0.0
                    
                    writer.writerow([sheet_name, sheet_name_val, db_id, db_name, competencia, base, bonus, comissao, descontos, total_val, "Importado auto"])
                else:
                    writer.writerow([sheet_name, sheet_name_val, "", "NÃO ENCONTRADO", competencia, total_val, 0, 0, 0, total_val, "Verificar Nome"])

    print(f"\nExtração concluída!")
    print(f"Arquivo gerado: {output_file}")
    print("Abra no Excel, confira as colunas (Ajuste o Mês/Ano e os valores de Base/Bônus/Glosas) e quando estiver ok, rodaremos o importador.")

if __name__ == "__main__":
    main()
