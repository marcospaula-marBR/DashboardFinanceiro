import urllib.request
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import shutil
import os

SUPABASE_URL = "https://ngtjhwswbbivqajtpjvg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5ndGpod3N3YmJpdnFhanRwanZnIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2OTE4MjM2MCwiZXhwIjoyMDg0NzU4MzYwfQ.2TPnOfnAzeWG23Y-VuDKxxzQ9QdbHwrnHdVBhS9hU28"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

def http(method, path, body=None):
    url = SUPABASE_URL + path
    data = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            text = r.read().decode()
            return r.status, json.loads(text) if text else []
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        return e.code, err

def main():
    print("=" * 80)
    print("CONSOLIDATION: Building BASE_UNICA_PEOPLE in Dianna.xlsx...")
    print("=" * 80)

    # 1. Fetch employees and loans
    print("Fetching employees...")
    status_e, emps = http("GET", "/rest/v1/employees?select=*")
    if status_e != 200:
        print("Error fetching employees:", emps)
        return
        
    print("Fetching loans...")
    status_l, loans = http("GET", "/rest/v1/employee_loans?select=*")
    if status_l != 200:
        print("Error fetching loans:", loans)
        return

    print(f"Loaded {len(emps)} employees and {len(loans)} loans.")

    # 2. Join employees and loans (LEFT JOIN)
    # Group loans by employee_id
    loans_by_emp = {}
    for l in loans:
        eid = l.get('employee_id')
        if eid not in loans_by_emp:
            loans_by_emp[eid] = []
        loans_by_emp[eid].append(l)

    joined_rows = []
    for e in emps:
        emp_loans = loans_by_emp.get(e['id'], [])
        if emp_loans:
            for l in emp_loans:
                joined_rows.append({
                    "emp": e,
                    "loan": l
                })
        else:
            joined_rows.append({
                "emp": e,
                "loan": {}
            })

    print(f"Total joined rows: {len(joined_rows)}")

    # 3. Open or create Dianna.xlsx
    dianna_path = "d:/DRE-V33-Dianna/dashboard-v2/public/Dianna.xlsx"
    desktop_path = "C:/Users/MarBrasil/OneDrive - Mar Brasil/Área de Trabalho/Dianna_Consolidada.xlsx"

    if os.path.exists(dianna_path):
        print(f"Loading existing Dianna.xlsx from {dianna_path}...")
        wb = openpyxl.load_workbook(dianna_path)
    else:
        print("Dianna.xlsx not found, creating new workbook...")
        wb = openpyxl.Workbook()

    # If BASE_UNICA_PEOPLE already exists, remove it to overwrite
    sheet_name = "BASE_UNICA_PEOPLE"
    if sheet_name in wb.sheetnames:
        print(f"Removing existing sheet '{sheet_name}'...")
        wb.remove(wb[sheet_name])

    # Create sheet at index 0
    ws = wb.create_sheet(title=sheet_name, index=0)

    # 4. Define headers and field mappings
    headers = [
        # Employee fields
        "ID_COLABORADOR_DO_NOT_CHANGE",
        "NOME_COMPLETO_RAZAO_SOCIAL",
        "EMPRESA_MarBR_OU_DZM",
        "TIPO_VINCULO_CLT_MEI_ESTAGIARIO_PJ",
        "NOME_SOCIO_RESPONSAVEL",
        "CPF_SOCIO_RESPONSAVEL",
        "CPF_CNPJ_DOCUMENTO",
        "CHAVE_PIX",
        "CARGO",
        "SETOR",
        "SALARIO_FIXO_BRUTO",
        "VALOR_BONUS_FIXO",
        "VALOR_COMISSAO_FIXA",
        "DATA_ADMISSAO_INICIO_AAAA_MM_DD",
        "DATA_DESLIGAMENTO_AAAA_MM_DD",
        "STATUS_Ativo_Ferias_Inativo_Provisao",
        "TELEFONE",
        "EMAIL",
        "CEP",
        "RUA",
        "NUMERO",
        "COMPLEMENTO",
        "BAIRRO",
        "CIDADE",
        "ESTADO_UF",
        "GENERO_M_F",
        "ESTADO_CIVIL",
        "TELEFONE_PROFISSIONAL",
        "EMAIL_PROFISSIONAL",
        # Loan fields
        "EMPRESTIMO_ID_DO_NOT_CHANGE",
        "EMPRESTIMO_VALOR_TOTAL",
        "EMPRESTIMO_PARCELAS_TOTAIS",
        "EMPRESTIMO_PARCELAS_PAGAS",
        "EMPRESTIMO_VALOR_PAGO_EXTRA",
        "EMPRESTIMO_MESES_ADIADOS",
        "EMPRESTIMO_MES_INICIO_AAAA_MM",
        "EMPRESTIMO_DATA_SOLICITACAO_AAAA_MM_DD",
        "EMPRESTIMO_OBSERVACOES",
        "EMPRESTIMO_LINK_CONTRATO"
    ]

    emp_mapping = [
        ("id", ""),
        ("full_name", ""),
        ("company", ""),
        ("employment_type", ""),
        ("responsible_name", ""),
        ("responsible_cpf", ""),
        ("document_id", ""),
        ("pix_key", ""),
        ("job_role", ""),
        ("department", ""),
        ("remuneration_fixed", 0),
        ("remuneration_bonus", 0),
        ("remuneration_commission", 0),
        ("start_date", ""),
        ("resignation_date", ""),
        ("status", ""),
        ("phone", ""),
        ("email", ""),
        ("zip_code", ""),
        ("street", ""),
        ("number", ""),
        ("complement", ""),
        ("neighborhood", ""),
        ("city", ""),
        ("state", ""),
        ("gender", ""),
        ("marital_status", ""),
        ("phone_professional", ""),
        ("email_professional", "")
    ]

    loan_mapping = [
        ("id", ""),
        ("amount", 0),
        ("installments", 0),
        ("paid_installments", 0),
        ("amount_paid_extra", 0),
        ("postponed_months", 0),
        ("start_cycle", ""),
        ("request_date", ""),
        ("notes", ""),
        ("contract_url", "")
    ]

    # Write headers
    ws.append(headers)

    # Styling colors
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # soft green for profile
    blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # soft blue for loan
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    border_side = Side(border_style="thin", color="CCCCCC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Format headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # First 29 columns are profile (green), next 10 are loans (blue)
        if col_idx <= 29:
            cell.fill = green_fill
        else:
            cell.fill = blue_fill
    
    ws.row_dimensions[1].height = 28

    # Write data
    print("Writing records...")
    for idx, row in enumerate(joined_rows):
        emp_data = row["emp"]
        loan_data = row["loan"]

        row_values = []
        for key, default in emp_mapping:
            val = emp_data.get(key)
            row_values.append(val if val is not None else default)

        for key, default in loan_mapping:
            val = loan_data.get(key)
            row_values.append(val if val is not None else default)

        ws.append(row_values)

        # Style data row
        row_num = idx + 2
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            if col_idx in (11, 12, 13, 31, 34): # numeric values
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (32, 33, 35): # integer counts
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Autofit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to Dianna.xlsx
    print(f"Saving changes to {dianna_path}...")
    wb.save(dianna_path)

    # Also save a copy on the user's Desktop as Dianna_Consolidada.xlsx
    print(f"Saving a copy directly to user Desktop: {desktop_path}...")
    wb.save(desktop_path)
    print("[SUCCESS] Consolidation finished! Sheet 'BASE_UNICA_PEOPLE' created as the first tab.")

if __name__ == "__main__":
    main()
