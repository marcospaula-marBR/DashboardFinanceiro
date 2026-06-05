import openpyxl

def main():
    path = "d:/DRE-V33-Dianna/dashboard-v2/public/Dianna.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = "CONSOLIDADO ATIVO INATIVO NOVO"
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        return
    
    sh = wb[sheet_name]
    print(f"Sheet {sheet_name}: {sh.max_row} rows, {sh.max_column} columns")
    rows = list(sh.iter_rows(max_row=10, values_only=True))
    
    print("\nHeaders:")
    print(rows[0])
    
    print("\nFirst 5 data rows:")
    for i in range(1, 6):
        if i < len(rows):
            print(f"Row {i+1}: {rows[i]}")

if __name__ == "__main__":
    main()
