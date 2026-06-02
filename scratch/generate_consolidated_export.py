import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import urllib.request
import json
import sys
import os

# Supabase Configurations
SUPABASE_URL = "https://ngtjhwswbbivqajtpjvg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5ndGpod3N3YmJpdnFhanRwanZnIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2OTE4MjM2MCwiZXhwIjoyMDg0NzU4MzYwfQ.2TPnOfnAzeWG23Y-VuDKxxzQ9QdbHwrnHdVBhS9hU28"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

def http(method, path):
    url = SUPABASE_URL + path
    req = urllib.request.Request(url, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            text = r.read().decode()
            return r.status, json.loads(text) if text else []
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        return e.code, err

def add_db_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)
    if not data:
        ws.append(["Nenhum dado encontrado"])
        return
    
    # Write header
    headers = list(data[0].keys())
    ws.append(headers)
    
    # Write rows
    for row in data:
        row_vals = []
        for h in headers:
            val = row[h]
            # Convert list or dict to JSON string for excel readability
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            row_vals.append(val)
        ws.append(row_vals)

def copy_dianna_sheets(src_wb, dest_wb):
    for name in src_wb.sheetnames:
        # Only copy key sheets to avoid making the excel file too bloated
        if name in ['CONSOLIDADO ATIVO INATIVO NOVO', 'MEI - NOVA', 'CLT NOVA']:
            src_ws = src_wb[name]
            dest_ws = dest_wb.create_sheet(title=f"Original_{name[:20]}")
            for row in src_ws.iter_rows(values_only=True):
                dest_ws.append(list(row))

def main():
    print("=" * 80)
    print("EXPORT: Generating Consolidated RH Spreadsheet")
    print("=" * 80)

    # 1. Fetch tables from Supabase
    print("\n[1/3] Buscando tabelas do Supabase...")
    
    print("  - Buscando 'employees'...")
    status, emps = http("GET", "/rest/v1/employees?select=*&order=full_name.asc")
    if status != 200:
        print(f"[ERROR] Falha ao carregar employees: {emps}")
        return
        
    print("  - Buscando 'people_employment_bonds'...")
    status, bonds = http("GET", "/rest/v1/people_employment_bonds?select=*&order=start_date.asc")
    
    print("  - Buscando 'people_monthly_costs'...")
    status, costs = http("GET", "/rest/v1/people_monthly_costs?select=*&order=competencia.asc")
    
    print("  - Buscando 'employee_loans'...")
    status, loans = http("GET", "/rest/v1/employee_loans?select=*&order=request_date.asc")
    
    print("  - Buscando 'employee_history'...")
    status, history = http("GET", "/rest/v1/employee_history?select=*&order=change_date.asc")

    # 2. Open source Dianna.xlsx
    dianna_path = "D:/DRE-V33-Dianna/dashboard-v2/public/Dianna.xlsx"
    print(f"\n[2/3] Carregando planilhas locais de: {dianna_path}")
    if not os.path.exists(dianna_path):
        print(f"[ERROR] Planilha Dianna.xlsx não encontrada em: {dianna_path}")
        return
    src_wb = openpyxl.load_workbook(dianna_path, data_only=True)

    # 3. Create dest workbook and copy sheets
    print("\n[3/3] Montando nova planilha consolidada...")
    dest_wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = dest_wb.active
    if default_sheet:
        dest_wb.remove(default_sheet)

    # Add Supabase tables
    print("  - Adicionando abas do Supabase...")
    add_db_sheet(dest_wb, "DB_Employees", emps)
    add_db_sheet(dest_wb, "DB_Vinculos", bonds)
    add_db_sheet(dest_wb, "DB_Custos_Mensais", costs)
    add_db_sheet(dest_wb, "DB_Emprestimos", loans)
    add_db_sheet(dest_wb, "DB_Historico_RH", history)

    # Add Dianna sheets
    print("  - Copiando abas originais do arquivo Dianna.xlsx...")
    copy_dianna_sheets(src_wb, dest_wb)

    # Save to public directory
    output_path = "D:/DRE-V33-Dianna/dashboard-v2/public/Dados_Consolidados_RH.xlsx"
    print(f"\nSalvando arquivo consolidado em: {output_path}")
    dest_wb.save(output_path)
    print("Planilha consolidada gravada com sucesso!")
    print("=" * 80)

if __name__ == "__main__":
    main()
