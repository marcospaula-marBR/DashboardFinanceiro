import openpyxl

def main():
    path = "d:/DRE-V33-Dianna/dashboard-v2/public/Dianna.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    print("Sheets in Dianna.xlsx:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
        # Get first few rows of each sheet
        sh = wb[sheet]
        rows = list(sh.iter_rows(max_row=3, values_only=True))
        if rows:
            print(f"    Headers: {rows[0][:10]}")
            if len(rows) > 1:
                print(f"    Row 2: {rows[1][:10]}")
        print()

if __name__ == "__main__":
    main()
